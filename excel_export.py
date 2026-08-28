"""Create a formatted Excel workbook from parsed receipts matching the accounting template."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from parser import Receipt

BOLD = Font(name="Calibri", bold=True, size=11)
HEADER_FONT = Font(name="Calibri", bold=True, size=11)
ITEM_FONT = Font(name="Segoe UI Emoji", size=11)
DISCOUNT_FONT = Font(name="Segoe UI Emoji", size=11)
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NUMBER_FORMAT = "#,##0"
BOTTOM_BORDER = Border(bottom=Side(style="thin", color="000000"))


def export(receipts: list[Receipt], output_path: str, sheet_title: str = "Hoá đơn") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 18

    current_row = 1
    for receipt in receipts:
        # Receipt Header (e.g. Receipt-GM-016.pdf)
        header = ws.cell(current_row, 1, receipt.display_name)
        header.font = HEADER_FONT
        current_row += 1

        first_amount_row = None
        for row in receipt.rows:
            if row.type == "qty":
                qty_cell = ws.cell(current_row, 1, row.label)
                qty_cell.font = BOLD
                current_row += 1
                continue

            label_cell = ws.cell(current_row, 1, row.label)
            label_cell.alignment = Alignment(wrap_text=True, vertical="top")

            if row.amount is not None:
                amount_cell = ws.cell(current_row, 2, row.amount)
                amount_cell.number_format = NUMBER_FORMAT
                amount_cell.alignment = Alignment(horizontal="right")
                if first_amount_row is None:
                    first_amount_row = current_row

            if row.type == "discount":
                label_cell.font = DISCOUNT_FONT
                if row.amount is not None:
                    amount_cell.font = DISCOUNT_FONT
            else:
                label_cell.font = ITEM_FONT
                if row.amount is not None:
                    amount_cell.font = ITEM_FONT

            current_row += 1

        # Total row (Tổng | =SUM(B...:B...))
        total_label = ws.cell(current_row, 1, "Tổng")
        total_label.font = BOLD
        total_label.fill = YELLOW
        total_label.border = BOTTOM_BORDER

        total_value = ws.cell(current_row, 2)
        if first_amount_row is not None:
            total_value.value = f"=SUM(B{first_amount_row}:B{current_row - 1})"
        else:
            total_value.value = 0

        total_value.font = BOLD
        total_value.fill = YELLOW
        total_value.border = BOTTOM_BORDER
        total_value.number_format = NUMBER_FORMAT
        total_value.alignment = Alignment(horizontal="right")

        current_row += 1

    ws.freeze_panes = "A1"
    wb.save(output_path)
    return output_path
